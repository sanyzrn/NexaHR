"""ورود دسته‌ای پرسنل از فایل Excel — تجزیه و اعتبارسنجی، بدون هیچ نوشتنی.

راه‌اندازی یک مشتری تازه یعنی وارد کردن ده‌ها یا صدها پرسنل. تایپ دستی هم کند
است هم مستعد خطا، و خطای تایپی در «کد پرسنلی» بعداً به یک پروندهٔ ارزیابی
گره‌خورده به فرد اشتباه تبدیل می‌شود.

دو تصمیم که شکل این ماژول را تعیین کرده‌اند:

۱. **اعتبارسنجی کاملاً از درج جداست.** این فایل هیچ‌وقت چیزی نمی‌نویسد. خروجی‌اش
   یک گزارش است: کدام ردیف‌ها سالم‌اند، کدام‌ها نیستند و دقیقاً چرا. UI همین را
   نشان می‌دهد و تا تأیید کاربر چیزی درج نمی‌شود — چون «۲۰۰ ردیف وارد شد و ۳تا
   اشتباه بود» را نمی‌شود به‌سادگی برگرداند.

۲. **ورودی باید همان چیزی را بپذیرد که خروجی تولید می‌کند.** جریان واقعی HR این
   است: خروجی اکسل بگیر، در اکسل ویرایش کن، دوباره وارد کن. پس ارقام فارسی،
   تاریخ شمسی و «بله/خیر» و «فعال/غیرفعال» همگی پذیرفته می‌شوند. یک تست همین
   رفت‌وبرگشت را می‌سنجد.
"""
from __future__ import annotations

from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from app.schemas.personnel import CreatedAccount

import re
from dataclasses import dataclass, field
from datetime import date
from io import BytesIO

import jdatetime
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.enums import PersonnelStatus, UserRole
from app.models.personnel import Personnel
from app.models.user import User

# همان حداقلی که فرم ساخت کاربر اعمال می‌کند؛ دو مسیر ورود نباید دو قانون داشته باشند.
from app.schemas.user import _PASSWORD_MIN_LENGTH as PASSWORD_MIN_LENGTH
from app.services.org_unit import join_site
from app.services.workflow import may_act_at

# همان ستون‌های build_personnel_workbook، به‌علاوهٔ یک ستون اختیاری برای نام کاربری
COLUMNS = [
    "کد پرسنلی",
    "نام و نام خانوادگی",
    "عنوان شغلی",
    # محل و واحد دو ستون جدا هستند، ولی در یک رشته ذخیره می‌شوند («محل / واحد»).
    #
    # چرا دو ستون: تا امروز فقط «واحد سازمانی» بود و کسی که فایل را پر می‌کرد
    # باید قرارداد جداکننده را می‌دانست. هر کس که نمی‌دانست، ردیفی می‌ساخت که در
    # هیچ فیلترِ محلی دیده نمی‌شد — بی‌آنکه خطایی بگیرد.
    #
    # ستون اختیاری است: فایل قدیمی که «کارخانه / فروش» را در همان یک ستون دارد،
    # همچنان درست خوانده می‌شود.
    "محل",
    "واحد سازمانی",
    "مدیر",
    "وضعیت",
    "شروع قرارداد",
    "پایان قرارداد",
    "نام کاربری",
    # رمز اولیه، اختیاری. اگر خالی بماند سامانه خودش یکی می‌سازد و *یک‌بار* در
    # گزارش پایانِ ایمپورت نشان می‌دهد. پرکردنش یعنی رمزها را از قبل خودتان
    # می‌دانید و لازم نیست از آن صفحه برداریدشان.
    #
    # هر دو حالت `must_change_password` را روشن می‌گذارند: رمزی که در یک فایل
    # اکسل نوشته شده، رمز نیست — یک بلیط ورود یک‌بارمصرف است.
    "رمز اولیه",
    # سه ستون زنجیرهٔ ارزیابی. هر سه اختیاری‌اند و با *نام* پر می‌شوند، نه با
    # شناسه: کسی که فایل پرسنلی را در اکسل پر می‌کند، id کاربر را نمی‌داند.
    #
    # بدون این‌ها، ایمپورت فقط پرسنل می‌ساخت و زنجیره خالی می‌ماند — یعنی
    # بلافاصله بعد از یک ایمپورت ۴۲ نفره، ۴۲ نفر داشتید که هیچ‌کس نمی‌توانست
    # ارزیابی‌شان کند، و تنظیمش ۴۲ بار باز کردن فرم ویرایش بود.
    "مسئول مستقیم",
    "معاونت مربوطه",
    "مدیرعامل",
]
#: «محل» عمداً اجباری نیست — سازمانی که یک محل بیشتر ندارد نباید ستونی پر کند
#: که هیچ‌چیز را تفکیک نمی‌کند.
REQUIRED_COLUMNS = [
    "کد پرسنلی",
    "نام و نام خانوادگی",
    "عنوان شغلی",
    "واحد سازمانی",
    "شروع قرارداد",
    "پایان قرارداد",
]

#: ستون‌های زنجیره، و مرحله‌ای که هرکدام پر می‌کنند.
CHAIN_COLUMNS: tuple[tuple[str, str, UserRole], ...] = (
    ("مسئول مستقیم", "unit_supervisor_user_id", UserRole.unit_supervisor),
    ("معاونت مربوطه", "deputy_user_id", UserRole.deputy),
    ("مدیرعامل", "ceo_user_id", UserRole.ceo),
)

#: مقادیری که یعنی «این مرحله را ندارد» — نه اینکه یادشان رفته پر کنند.
_ABSENT_WORDS = {"", "-", "—", "ندارد", "نامشخص"}

# ارقام فارسی و عربی → اسکی. بدون این، «۱۴۰۵/۰۱/۰۱» که خودِ ما تولید کرده‌ایم
# هنگام بازگشت غیرقابل تجزیه می‌شد.
_DIGITS = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")

_TRUE_WORDS = {"بله", "بلی", "آری", "true", "yes", "1", "y"}
_FALSE_WORDS = {"خیر", "نه", "false", "no", "0", "n", ""}
_ACTIVE_WORDS = {"فعال", "active", "1", ""}
_INACTIVE_WORDS = {"غیرفعال", "غیر فعال", "inactive", "0"}

_USERNAME_RE = re.compile(r"^[a-zA-Z0-9._-]{3,50}$")
# سال شمسی تا ۱۵۰۰ و میلادی از ۱۹۰۰ به بعد — برای تشخیص تقویم از روی خود عدد
_JALALI_MAX_YEAR = 1500


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        # نیم‌فاصله و فاصله‌های نامرئی اکسل، مقایسهٔ رشته‌ای را بی‌صدا خراب می‌کنند
        return value.replace("‌", "‌").replace("\xa0", " ").strip()
    return str(value).strip()


def _normalise_digits(value: str) -> str:
    return value.translate(_DIGITS)


def parse_flexible_date(raw: object) -> date | None:
    """تاریخ شمسی یا میلادی، با ارقام فارسی یا اسکی، یا تاریخِ خودِ اکسل.

    سال تشخیص تقویم را می‌دهد: چیزی حوالی ۱۴۰۵ شمسی است و حوالی ۲۰۲۶ میلادی.
    این ابهام واقعی ندارد چون بازه‌ها اصلاً هم‌پوشانی ندارند.
    """
    if raw is None or _text(raw) == "":
        return None
    # اکسل ممکن است سلول را از قبل به datetime تبدیل کرده باشد
    if isinstance(raw, date):
        return raw
    if hasattr(raw, "date"):
        return raw.date()

    cleaned = _normalise_digits(_text(raw)).replace("-", "/").replace(".", "/")
    parts = [p for p in cleaned.split("/") if p]
    if len(parts) != 3:
        return None
    try:
        year, month, day = (int(p) for p in parts)
    except ValueError:
        return None

    try:
        if year <= _JALALI_MAX_YEAR:
            return jdatetime.date(year, month, day).togregorian()
        return date(year, month, day)
    except ValueError:
        return None


def _parse_bool(raw: object) -> bool | None:
    value = _normalise_digits(_text(raw)).lower()
    if value in _TRUE_WORDS:
        return True
    if value in _FALSE_WORDS:
        return False
    return None


def _parse_status(raw: object) -> PersonnelStatus | None:
    value = _normalise_digits(_text(raw)).lower()
    if value in _ACTIVE_WORDS:
        return PersonnelStatus.active
    if value in _INACTIVE_WORDS:
        return PersonnelStatus.inactive
    return None


@dataclass
class ImportRow:
    """یک ردیف فایل، همراه با هرچه اشکال دارد.

    `row_number` شمارهٔ ردیف در خود اکسل است (با احتساب سرستون) تا کاربر بتواند
    مستقیم برود همان‌جا؛ «ردیف سوم دادهٔ معتبر» به درد کسی نمی‌خورد.
    """

    row_number: int
    personnel_code: str = ""
    full_name: str = ""
    job_title: str = ""
    org_unit: str = ""
    is_manager: bool = False
    status: PersonnelStatus = PersonnelStatus.active
    contract_start_date: date | None = None
    contract_end_date: date | None = None
    username: str | None = None
    #: رمزِ دادهٔ کاربر. None یعنی سامانه خودش می‌سازد.
    initial_password: str | None = None
    # شناسهٔ کاربرِ هر مرحله، پس از تطبیق نام. None یعنی یا ستون خالی بوده یا
    # نامش پیدا نشد — که دومی خودش یک خطای ردیف است، پس این دو با هم قاطی
    # نمی‌شوند.
    unit_supervisor_user_id: int | None = None
    deputy_user_id: int | None = None
    ceo_user_id: int | None = None
    errors: list[str] = field(default_factory=list)

    @property
    def has_chain(self) -> bool:
        """آیا آن‌قدر زنجیره دارد که بشود ساختش؟

        مدیرعامل الزامی است (کسی باید پرونده را ببندد) و دست‌کم یکی از دو مرحلهٔ
        میانی، وگرنه هیچ‌کس نمره نمی‌دهد.
        """
        return self.ceo_user_id is not None and (
            self.unit_supervisor_user_id is not None or self.deputy_user_id is not None
        )

    @property
    def ok(self) -> bool:
        return not self.errors


@dataclass
class ImportPreview:
    rows: list[ImportRow]
    """خطاهای مربوط به کل فایل (ستون جاافتاده، فایل خالی) — نه یک ردیف خاص."""
    file_errors: list[str] = field(default_factory=list)

    @property
    def valid(self) -> list[ImportRow]:
        return [r for r in self.rows if r.ok]

    @property
    def invalid(self) -> list[ImportRow]:
        return [r for r in self.rows if not r.ok]


def _read_sheet(content: bytes) -> tuple[list[str], list[tuple[int, tuple]]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return [], []
    header = [_text(c) for c in rows[0]]
    body = [(i + 2, r) for i, r in enumerate(rows[1:])]
    return header, body


def _evaluator_index(db: Session) -> tuple[dict[str, list[User]], dict[str, User]]:
    """کاربرانِ فعال، یک‌بار، برای تطبیق نام‌های ستون‌های زنجیره.

    هم با نام و هم با نام کاربری تطبیق داده می‌شود: فایلِ منابع انسانی معمولاً
    نام فارسی دارد، ولی کسی که فایل دوم را می‌سازد ممکن است نام کاربری بنویسد،
    و ردکردنِ آن فقط یک مانع بی‌دلیل است.

    نام‌های هم‌نام در یک فهرست جمع می‌شوند نه اینکه یکی دیگری را بپوشاند —
    انتخابِ خاموشِ یکی از دو «محمد محمدی»، همان کسی است که بعداً پای تأیید
    پروندهٔ اشتباهی می‌نشیند.
    """
    users = list(db.scalars(select(User).where(User.is_active.is_(True))))
    by_name: dict[str, list[User]] = {}
    for user in users:
        if user.full_name:
            by_name.setdefault(user.full_name.strip(), []).append(user)
    return by_name, {u.username: u for u in users}


def _resolve_chain(
    item: ImportRow,
    raw: tuple,
    cell,
    by_name: dict[str, list[User]],
    by_username: dict[str, User],
    sole_ceo: User | None,
) -> None:
    """نام‌های ستون‌های زنجیره را به کاربر تبدیل می‌کند.

    نامی که پیدا نشود یک خطای ردیف است، نه یک هشدار. اگر بی‌صدا رد می‌شد،
    ایمپورت «موفق» گزارش می‌داد و آن پرسنل بدون زنجیره می‌ماند — یعنی همان
    وضعیتی که این ستون‌ها برای رفعش اضافه شدند، فقط این بار پنهان.
    """
    for column, attribute, stage_role in CHAIN_COLUMNS:
        written = _text(cell(raw, column))
        if written in _ABSENT_WORDS:
            continue

        # هیچ‌کس ارزیابِ خودش نیست (P0-10). پیش از تطبیق نام سنجیده می‌شود،
        # وگرنه وقتی آن فرد هنوز حساب کاربری ندارد، پیامِ «پیدا نشد» جای
        # تشخیصِ درست را می‌گرفت — و کاربر می‌رفت حسابی بسازد که مشکل را حل
        # نمی‌کرد.
        #
        # مسیرهای دیگر این را با `ensure_evaluators_are_not_the_subject`
        # می‌سنجند، ولی آن از روی `users.personnel_id` کار می‌کند و در لحظهٔ
        # پیش‌نمایش این پرسنل هنوز ساخته نشده. تریگرهای دیتابیس (مایگریشن
        # c3e8b1a76d94) پشتیبان نهایی‌اند، ولی وسط ایمپورت یک خطای پایگاه‌داده
        # می‌دهند نه پیامی که بشود قبلش خواند.
        if written == item.full_name:
            item.errors.append(f"«{column}»: یک نفر نمی‌تواند ارزیابِ خودش باشد")
            continue

        matches = by_name.get(written)
        if not matches:
            from_username = by_username.get(written)
            matches = [from_username] if from_username else []

        if not matches:
            item.errors.append(
                f"«{column}»: کاربری با نام «{written}» پیدا نشد؛ اول حسابش را بسازید"
            )
            continue
        if len(matches) > 1:
            # انتخابِ خاموشِ یکی از دو هم‌نام، همان کسی است که بعداً پای تأیید
            # پروندهٔ اشتباهی می‌نشیند.
            item.errors.append(
                f"«{column}»: بیش از یک کاربر با نام «{written}» هست؛ به‌جای نام، نام کاربری را بنویسید"
            )
            continue

        user = matches[0]
        if not may_act_at(user.role, stage_role):
            item.errors.append(
                f"«{column}»: «{written}» نمی‌تواند در این مرحله قرار بگیرد"
            )
            continue
        setattr(item, attribute, user.id)

    # ستون مدیرعامل معمولاً در فایل‌های واقعی خالی است، چون سازمان یکی بیشتر
    # ندارد و تکرارش در هر ردیف کار بیهوده‌ای است.
    if item.ceo_user_id is None and sole_ceo is not None:
        item.ceo_user_id = sole_ceo.id


    # پرسنلِ «مدیر» مسئول واحد ندارد — همان قانونی که فرم دسترسی هم اعمال می‌کند.
    if item.is_manager:
        item.unit_supervisor_user_id = None

    # صندلیِ تکراری. فقط جفت‌هایی رد می‌شوند که بیان درست‌تری دارند — «مسئول
    # مستقیم = مدیرعامل» عمداً مجاز است، چون تنها راه ثبتِ کسی است که مستقیم زیر
    # نظر مدیرعامل کار می‌کند (توضیح کامل در services/self_evaluation.py).
    redundant = [
        (
            item.unit_supervisor_user_id,
            item.deputy_user_id,
            "«مسئول مستقیم» و «معاونت مربوطه» یک نفرند؛ اگر این فرد مستقیماً توسط "
            "معاونت ارزیابی می‌شود، «مسئول مستقیم» را خالی بگذارید",
        ),
        (
            item.deputy_user_id,
            item.ceo_user_id,
            "«معاونت مربوطه» و «مدیرعامل» یک نفرند؛ اگر معاونتی بالای سر این فرد "
            "نیست، «معاونت مربوطه» را خالی بگذارید",
        ),
    ]
    for first, second, message in redundant:
        if first is not None and first == second:
            item.errors.append(message)
            break


def parse_workbook(content: bytes, db: Session) -> ImportPreview:
    """فایل را می‌خواند، هر ردیف را اعتبارسنجی می‌کند و گزارش می‌دهد. چیزی نمی‌نویسد."""
    try:
        header, body = _read_sheet(content)
    except Exception:  # noqa: BLE001 — هر خرابی فایل، یک پیام کاربرپسند می‌شود
        return ImportPreview(rows=[], file_errors=["فایل خوانده نشد؛ باید یک فایل معتبر Excel (.xlsx) باشد."])

    if not header:
        return ImportPreview(rows=[], file_errors=["فایل خالی است."])

    index = {name: i for i, name in enumerate(header)}
    missing = [c for c in REQUIRED_COLUMNS if c not in index]
    if missing:
        return ImportPreview(
            rows=[],
            file_errors=[f"این ستون‌ها در فایل نیستند: {'، '.join(missing)}"],
        )

    def cell(row: tuple, column: str) -> object:
        position = index.get(column)
        if position is None or position >= len(row):
            return None
        return row[position]

    # یک‌بار خوانده می‌شوند تا برای هر ردیف یک کوئری جدا نزنیم
    existing_codes = {c for (c,) in db.execute(select(Personnel.personnel_code))}
    existing_usernames = {u for (u,) in db.execute(select(User.username))}
    by_name, by_username = _evaluator_index(db)
    # اگر سازمان دقیقاً یک مدیرعامل فعال دارد، ستون خالیِ «مدیرعامل» همان را
    # می‌گیرد. نوشتنِ یک نام تکراری در ۴۲ ردیف، کاری است که فایل می‌تواند
    # نکند — و وقتی گزینه یکی است، حدسی در کار نیست.
    ceos = [u for users in by_name.values() for u in users if u.role is UserRole.ceo]
    sole_ceo = ceos[0] if len(ceos) == 1 else None

    seen_codes: dict[str, int] = {}
    seen_usernames: dict[str, int] = {}
    parsed: list[ImportRow] = []

    for number, raw in body:
        # ردیف کاملاً خالی (دنبالهٔ سطرهای خالی انتهای فایل) نه خطاست نه داده
        if all(_text(c) == "" for c in raw):
            continue

        item = ImportRow(row_number=number)
        item.personnel_code = _normalise_digits(_text(cell(raw, "کد پرسنلی")))
        item.full_name = _text(cell(raw, "نام و نام خانوادگی"))
        item.job_title = _text(cell(raw, "عنوان شغلی"))
        # اگر ستون «محل» پر باشد، به واحد چسبانده می‌شود؛ وگرنه هرچه در «واحد
        # سازمانی» نوشته شده همان است — که ممکن است خودش «کارخانه / فروش» باشد.
        item.org_unit = join_site(_text(cell(raw, "محل")), _text(cell(raw, "واحد سازمانی")))

        for label, value in (
            ("کد پرسنلی", item.personnel_code),
            ("نام و نام خانوادگی", item.full_name),
            ("عنوان شغلی", item.job_title),
            ("واحد سازمانی", item.org_unit),
        ):
            if not value:
                item.errors.append(f"«{label}» خالی است")

        if item.personnel_code:
            if item.personnel_code in existing_codes:
                item.errors.append("کد پرسنلی از قبل در سامانه ثبت شده است")
            elif item.personnel_code in seen_codes:
                item.errors.append(
                    f"کد پرسنلی تکراری است (ردیف {seen_codes[item.personnel_code]} همین فایل)"
                )
            else:
                seen_codes[item.personnel_code] = number

        is_manager = _parse_bool(cell(raw, "مدیر"))
        if is_manager is None:
            item.errors.append("ستون «مدیر» باید «بله» یا «خیر» باشد")
        else:
            item.is_manager = is_manager

        status = _parse_status(cell(raw, "وضعیت"))
        if status is None:
            item.errors.append("ستون «وضعیت» باید «فعال» یا «غیرفعال» باشد")
        else:
            item.status = status

        item.contract_start_date = parse_flexible_date(cell(raw, "شروع قرارداد"))
        item.contract_end_date = parse_flexible_date(cell(raw, "پایان قرارداد"))
        if item.contract_start_date is None:
            item.errors.append("«شروع قرارداد» خوانده نشد (نمونهٔ درست: ۱۴۰۵/۰۱/۰۱)")
        if item.contract_end_date is None:
            item.errors.append("«پایان قرارداد» خوانده نشد (نمونهٔ درست: ۱۴۰۶/۰۱/۰۱)")
        if (
            item.contract_start_date
            and item.contract_end_date
            and item.contract_end_date <= item.contract_start_date
        ):
            item.errors.append("«پایان قرارداد» باید بعد از «شروع قرارداد» باشد")

        username = _text(cell(raw, "نام کاربری"))
        if username:
            if not _USERNAME_RE.match(username):
                item.errors.append(
                    "نام کاربری فقط می‌تواند حروف انگلیسی، رقم، نقطه، خط تیره و زیرخط داشته باشد (۳ تا ۵۰ نویسه)"
                )
            elif username in existing_usernames:
                item.errors.append("این نام کاربری از قبل وجود دارد")
            elif username in seen_usernames:
                item.errors.append(
                    f"نام کاربری تکراری است (ردیف {seen_usernames[username]} همین فایل)"
                )
            else:
                seen_usernames[username] = number
                item.username = username

        # رمز فقط وقتی معنا دارد که حسابی در کار باشد.
        given_password = _text(cell(raw, "رمز اولیه"))
        if given_password and not username:
            item.errors.append("«رمز اولیه» بدون «نام کاربری» معنا ندارد")
        elif given_password:
            if len(given_password) < PASSWORD_MIN_LENGTH:
                item.errors.append(
                    f"«رمز اولیه» باید دست‌کم {PASSWORD_MIN_LENGTH} نویسه باشد"
                )
            else:
                item.initial_password = given_password

        _resolve_chain(item, raw, cell, by_name, by_username, sole_ceo)

        parsed.append(item)

    if not parsed:
        return ImportPreview(rows=[], file_errors=["فایل هیچ ردیف داده‌ای ندارد."])
    return ImportPreview(rows=parsed)


def build_template() -> bytes:
    """فایل نمونهٔ خالی با همان ستون‌ها و یک ردیف راهنما.

    بدون این، کاربر باید ستون‌ها را حدس بزند و اولین تلاشش تقریباً همیشه به خطای
    «ستون جاافتاده» می‌خورد.
    """
    from app.services.excel import _new_sheet, _to_bytes

    workbook, sheet = _new_sheet("پرسنل", COLUMNS)
    sheet.append(
        [
            # عمداً از بازهٔ کدهای دموی seed (P-1001 تا P-1003) فاصله دارد: وقتی
            # نمونه با یک پرسنل واقعی هم‌کد شود، همان فایلی که قرار است راهنما
            # باشد با خطای «کد پرسنلی تکراری است» رد می‌شود.
            "P-0000",
            "نام نمونه",
            "کارشناس",
            # یکی از سه محل. خالی‌گذاشتنش هم درست است — یعنی سازمان یک محل
            # بیشتر ندارد.
            "دفتر مرکزی",
            "واحد نمونه",
            "خیر",
            "فعال",
            "۱۴۰۵/۰۱/۰۱",
            "۱۴۰۶/۰۱/۰۱",
            "namuneh",
        ]
    )
    return _to_bytes(workbook)


# ── درج نهایی ─────────────────────────────────────────────────────────────
#
# این منطق قبلاً داخل روتِر `personnel.py` بود. دستیار هوشمند هم به همان مسیر
# درج نیاز دارد — و دو نسخه‌کردنِ «ساخت پرسنل + زنجیره + حساب» یعنی روزی یکی
# قاعدهٔ تازه‌ای بگیرد و دیگری نگیرد. پس منطق به سرویس منتقل شد و روتِر هم از
# همین تابع استفاده می‌کند؛ رفتارشان بایت‌به‌بایت یکی است.


def commit_import(db: Session, preview: ImportPreview, actor_user_id: int) -> ImportCommitResult:
    """درج ردیف‌های معتبرِ یک پیش‌نمایش، همه در یک تراکنش.

    فراخواننده مسئول است: پیش از این، `parse_workbook` تازه اجرا شده باشد و
    `file_errors` خالی باشد. ردیف‌های خطادار درج نمی‌شوند و در نتیجه شمرده
    می‌شوند — نه «همه یا هیچ»، چون یک غلط تایپی در ردیف ۱۹۰ نباید ۱۸۹ ردیف
    درستِ قبلی را دور بریزد.
    """
    from app.core.security import hash_password
    from app.models.evaluation_access import EvaluationAccess
    from app.schemas.personnel import CreatedAccount
    from app.services.audit import log_event
    from app.services.security_tokens import generate_temp_password

    accounts: list[CreatedAccount] = []
    created_personnel = 0
    chains_created = 0

    for row in preview.valid:
        personnel = Personnel(
            personnel_code=row.personnel_code,
            full_name=row.full_name,
            job_title=row.job_title,
            org_unit=row.org_unit,
            is_manager=row.is_manager,
            status=row.status,
            contract_start_date=row.contract_start_date,
            contract_end_date=row.contract_end_date,
            created_by_user_id=actor_user_id,
        )
        db.add(personnel)
        db.flush()
        created_personnel += 1
        log_event(
            db,
            actor_user_id=actor_user_id,
            event_type="personnel_created",
            new_value={
                "id": personnel.id,
                "personnel_code": personnel.personnel_code,
                "full_name": personnel.full_name,
                "imported": True,
            },
        )

        if row.has_chain:
            db.add(
                EvaluationAccess(
                    personnel_id=personnel.id,
                    unit_supervisor_user_id=row.unit_supervisor_user_id,
                    deputy_user_id=row.deputy_user_id,
                    ceo_user_id=row.ceo_user_id,
                    updated_by_user_id=actor_user_id,
                )
            )
            chains_created += 1
            log_event(
                db,
                actor_user_id=actor_user_id,
                event_type="evaluation_access_set",
                new_value={
                    "personnel_id": personnel.id,
                    "unit_supervisor_user_id": row.unit_supervisor_user_id,
                    "deputy_user_id": row.deputy_user_id,
                    "ceo_user_id": row.ceo_user_id,
                    "imported": True,
                },
            )

        if row.username:
            password = row.initial_password or generate_temp_password()
            user = User(
                username=row.username,
                password_hash=hash_password(password),
                role=UserRole.employee,
                personnel_id=personnel.id,
                is_active=True,
                must_change_password=True,
            )
            db.add(user)
            db.flush()
            # رمز عمداً در لاگ ممیزی نیست: لاگ ماندگار است و رمز نباید ماندگار شود.
            log_event(
                db,
                actor_user_id=actor_user_id,
                event_type="user_created",
                new_value={
                    "id": user.id,
                    "username": user.username,
                    "role": user.role.value,
                    "created_with_personnel": personnel.id,
                    "imported": True,
                },
            )
            accounts.append(
                CreatedAccount(
                    personnel_code=personnel.personnel_code,
                    full_name=personnel.full_name,
                    username=user.username,
                    temporary_password=password,
                )
            )

    log_event(
        db,
        actor_user_id=actor_user_id,
        event_type="personnel_imported",
        new_value={
            "created_personnel": created_personnel,
            "created_accounts": len(accounts),
            "skipped_rows": len(preview.invalid),
        },
    )
    db.commit()

    return ImportCommitResult(
        created_personnel=created_personnel,
        created_chains=chains_created,
        skipped_rows=len(preview.invalid),
        accounts=accounts,
    )


@dataclass
class ImportCommitResult:
    """نتیجهٔ درج: چند پرسنل، چند زنجیره، چند حساب؛ و رمزهای موقتِ یک‌بارمصرف."""

    created_personnel: int
    created_chains: int
    skipped_rows: int
    accounts: list[CreatedAccount]
