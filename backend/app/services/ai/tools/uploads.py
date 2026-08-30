"""فایل‌ها در گفت‌وگو: بارگذاری، بازرسی، اصلاح، ورود گروهی.

جریانِ کاملِ «اکسل ناقص» همان چیزی است که این ماژول برایش ساخته شده:

۱. کاربر فایل را در گفت‌وگو می‌اندازد → سرور آن را *مرحله‌بندی* می‌کند:
   `parse_workbook` رسمی اجرا می‌شود و گزارشِ خطای هر ردیف کنار بایت‌های خام
   ذخیره می‌گردد. هنوز هیچ ردیفی ساخته نشده.
۲. مدل با ``inspect_upload`` خطاها را می‌بیند و به زبان آدم توضیح می‌دهد؛
   اگر مقداری کم است، *می‌پرسد* — حدس نمی‌زند.
۳. جواب‌های کاربر با ``patch_upload_rows`` روی یک لایهٔ نازک (overlay) می‌نشیند
   و فایل دوباره با همان اعتبارسنجیِ رسمی خوانده می‌شود. نتیجه: دستیار و فرمِ
   دستیِ ورود فایل هر دو دقیقاً یک قانون می‌بینند؛ «اعتبارسنجیِ دومِ ضعیف‌تر»
   در کار نیست.
۴. وقتی همهٔ ردیف‌ها سالم شدند، ``import_personnel`` کنشِ «پرخطر» است: مدل
   فقط پیشنهادش را ثبت می‌کند و اجرا با تأییدِ صریح کاربر از نقطهٔ دیگر رخ
   می‌دهد — آن‌جا برای بار دوم اعتبارسنجی و مجوز سنجیده می‌شود.
"""
from __future__ import annotations

import hashlib
import json
from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.enums import Capability, UserRole
from app.schemas.auth import CurrentUser
from app.services.ai.tools.base import ToolContext, ToolOutcome, json_content, tool
from app.services.personnel_import import ImportPreview, parse_workbook

#: زندگیِ یک پیشنهادِ در انتظارِ تأیید. کنشِ سه‌روزپیش ممکن است دیگر مجاز
#: نباشد؛ بی‌انقضا، دکمهٔ قدیمی یک مینِ خاموش است.
PENDING_TTL_HOURS = 24

_ALLOWED_EXTENSIONS = (".xlsx", ".xlsm")


# ── مرحله‌بندی و لایه ──────────────────────────────────────────────────────


def structure_of(upload) -> dict:
    try:
        return json.loads(upload.structure_json or "{}")
    except ValueError:
        return {}


def is_personnel_staging(upload) -> bool:
    return structure_of(upload).get("kind") == "personnel_import"


def apply_overlay(content: bytes, overlay: dict) -> bytes:
    """ویرایش‌های کاربر روی خود فایل می‌نشیند؛ نه کنارش.

    کلیدِ لایه شمارهٔ ردیف اکسل (با احتساب سرستون) است و مقدارش نگاشتِ
    «نام ستون → مقدار». سلولِ درست پیدا می‌شود، مقدار *به‌صورت متن* نوشته
    می‌شود — همان چیزی که `parse_workbook` برای پاک‌سازی و تجزیهٔ تاریخِ
    شمسی و ارقام فارسی انتظار دارد.
    """
    if not overlay:
        return content
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active
    header: dict[str, int] = {}
    for cell in sheet[1]:
        label = str(cell.value or "").strip()
        if label and label not in header:
            header[label] = cell.column
    for row_number, fields in overlay.items():
        try:
            row_index = int(row_number)
        except (TypeError, ValueError):
            continue
        if not isinstance(fields, dict):
            continue
        for column, value in fields.items():
            position = header.get(str(column).strip())
            if position is None:
                continue
            sheet.cell(row=row_index, column=position, value=str(value if value is not None else ""))
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def revalidate(db: Session, upload) -> ImportPreview:
    """همان `parse_workbook` رسمی، روی فایل + لایه — بدون هیچ میان‌بُری."""
    structure = structure_of(upload)
    overlay = structure.get("overlay") or {}
    content = apply_overlay(upload.content, overlay)
    return parse_workbook(content, db)


def preview_summary(db: Session, upload) -> dict:
    """خلاصهٔ وضعیت مرحله‌بندی برای رابط و برای مدل."""
    preview = revalidate(db, upload)
    structure = structure_of(upload)
    return {
        "upload_id": upload.id,
        "filename": upload.filename,
        "kind": structure.get("kind"),
        "total_rows": len(preview.rows),
        "valid_count": len(preview.valid),
        "invalid_count": len(preview.invalid),
        "file_errors": list(preview.file_errors),
        "committed": bool(structure.get("committed")),
        "rows": _row_report(preview),
    }


def _row_report(preview: ImportPreview, only_problems: bool = False) -> list[dict]:
    rows = preview.invalid if only_problems else preview.rows
    out = []
    for row in rows:
        out.append(
            {
                "row_number": row.row_number,
                "personnel_code": row.personnel_code,
                "full_name": row.full_name,
                "org_unit": row.org_unit,
                "contract_start_date": row.contract_start_date.isoformat() if row.contract_start_date else None,
                "contract_end_date": row.contract_end_date.isoformat() if row.contract_end_date else None,
                "username": row.username,
                "has_account": bool(row.username),
                "has_chain": row.has_chain,
                "errors": list(row.errors),
            }
        )
    return out


def stage_upload(
    db: Session, user: CurrentUser, conversation_id: int, filename: str, mime_type: str, content: bytes
) -> tuple[object, dict]:
    """ذخیرهٔ فایل و ساختِ گزارشِ اولیه. هیچ ردیفی از پرسنل ساخته نمی‌شود.

    خروجی دوم، کارتِ رابط است: چند ردیف، چند خطا، و خلاصهٔ خطاها.
    """
    from app.models.ai import AiUpload

    lowered = (filename or "").lower()
    structure: dict = {"kind": "file"}
    if lowered.endswith(_ALLOWED_EXTENSIONS):
        preview = parse_workbook(content, db)
        if not preview.file_errors:
            structure = {
                "kind": "personnel_import",
                "overlay": {},
                "committed": False,
            }
        else:
            # فایل اکسل است ولی قالبِ پرسنل نیست: بازرسیِ عمومی می‌ماند.
            structure = {"kind": "excel", "file_errors": list(preview.file_errors)}
    else:
        structure = {
            "kind": "file",
            "note": "فقط فایل اکسل پرسنل (.xlsx) برای ورود گروهی پشتیبانی می‌شود",
        }

    upload = AiUpload(
        conversation_id=conversation_id,
        user_id=user.id,
        filename=(filename or "file")[:255],
        mime_type=(mime_type or "application/octet-stream")[:120],
        size_bytes=len(content),
        content_sha256=hashlib.sha256(content).hexdigest(),
        content=content,
        structure_json=json.dumps(structure, ensure_ascii=False),
    )
    db.add(upload)
    db.flush()
    summary = preview_summary(db, upload)
    # شمارنده‌ها و خطاها در خودِ ساختار می‌نشینند تا کارتِ رابط بدون بازرسیِ
    # دوباره، وضعیتِ آخر را نشان بدهد.
    structure.update({
        "total_rows": summary["total_rows"],
        "valid_count": summary["valid_count"],
        "invalid_count": summary["invalid_count"],
        "file_errors": summary["file_errors"],
    })
    upload.structure_json = json.dumps(structure, ensure_ascii=False)
    db.add(upload)
    db.flush()
    return upload, summary


def enrich_excel_structure(db: Session, upload) -> dict:
    """برای اکسلِ غیرِ پرسنلی: سرستون‌ها و چند ردیف اول، تا مدل بتواند توضیحش دهد."""
    try:
        workbook = load_workbook(BytesIO(upload.content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
    except Exception:  # noqa: BLE001 — فایل خراب، بازرسیِ عمومی هم خطای تمیز می‌دهد
        return {"error": "فایل خوانده نشد"}
    if not rows:
        return {"empty": True}
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    sample = []
    for raw in rows[1:6]:
        sample.append(
            {
                headers[i] if i < len(headers) and headers[i] else f"ستون {i + 1}": str(c) if c is not None else ""
                for i, c in enumerate(raw)
            }
        )
    return {
        "kind": "excel",
        "headers": headers,
        "row_count": len(rows) - 1,
        "sample": sample,
    }


def _upload_or_404(ctx: ToolContext, upload_id: int):
    from app.models.ai import AiUpload

    upload = ctx.db.get(AiUpload, int(upload_id))
    if upload is None or upload.user_id != ctx.user.id or upload.conversation_id != ctx.conversation_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "فایلی با این شناسه در این گفت‌وگو نیست")
    return upload


def _require_importer(ctx: ToolContext) -> None:
    if not (
        Capability.manage_personnel in ctx.caps
        or ctx.user.role == UserRole.hr
    ):
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            "ورود گروهی پرسنل فقط با مجوز «پرسنل و زنجیرهٔ ارزیابی» ممکن است",
        )


# ── ابزارها ────────────────────────────────────────────────────────────────


@tool(
    name="inspect_upload",
    description=(
        "بازرسی فایل بارگذاری‌شده در همین گفت‌وگو: ردیف‌ها، خطای هر ردیف و علتش. برای اکسل پرسنل، گزارش ورود گروهی است."
    ),
    category="فایل‌ها",
    read_only=True,
    parameters={"type": "object", "properties": {"upload_id": {"type": "integer"}}, "required": ["upload_id"]},
)
def inspect_upload(ctx: ToolContext, upload_id: int) -> ToolOutcome:
    upload = _upload_or_404(ctx, upload_id)
    if is_personnel_staging(upload):
        summary = preview_summary(ctx.db, upload)
        return ToolOutcome(
            content=json_content(summary),
            ui={"kind": "upload_report", "summary": {k: v for k, v in summary.items() if k != "rows"}},
            summary=(
                f"بازرسی «{upload.filename}»: "
                f"{summary['valid_count']} ردیف سالم، {summary['invalid_count']} خطادار"
            ),
        )
    structure = structure_of(upload)
    if structure.get("kind") in ("excel", "file"):
        payload = enrich_excel_structure(ctx.db, upload) if structure.get("kind") == "excel" else structure
        return ToolOutcome(
            content=json_content(payload),
            summary=f"بازرسی «{upload.filename}»",
        )
    return ToolOutcome(content=json_content(structure), summary=f"بازرسی «{upload.filename}»")


@tool(
    name="patch_upload_rows",
    description="اصلاح ردیف‌های اکسل پرسنلِ مرحله‌بندی‌شده با مقادیری که کاربر داده — "
    "مثلاً تکمیل تاریخ پایان قراردادِ جاافتاده. تغییرِ داده است؛ پس از تأیید کاربر "
    "اعمال و اعتبارسنجیِ رسمی از نو اجرا می‌شود.",
    category="فایل‌ها",
    read_only=False,
    risky=True,
    capabilities=(Capability.manage_personnel,),
    roles=(UserRole.hr,),
    parameters={
        "type": "object",
        "properties": {
            "upload_id": {"type": "integer"},
            "edits": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "row_number": {"type": "integer"},
                        "fields": {"type": "object", "description": "نگاشت نامِ ستون به مقدار تازه"},
                    },
                    "required": ["row_number", "fields"],
                },
                "description": 'نمونه: [{"row_number": 3, "fields": {"پایان قرارداد": "1406/05/01"}}]',
            },
        },
        "required": ["upload_id", "edits"],
    },
)
def patch_upload_rows(ctx: ToolContext, upload_id: int, edits: list) -> ToolOutcome:

    upload = _upload_or_404(ctx, upload_id)
    if not is_personnel_staging(upload):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "اصلاح ردیف فقط برای اکسل پرسنل معنا دارد")
    structure = structure_of(upload)
    if structure.get("committed"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "این فایل قبلاً وارد سامانه شده است")

    overlay: dict = structure.get("overlay") or {}
    applied = 0
    unknown_columns: list[str] = []
    for edit in edits or []:
        if not isinstance(edit, dict):
            continue
        row_number = str(edit.get("row_number"))
        fields = edit.get("fields")
        if not row_number or not isinstance(fields, dict):
            continue
        # نام ستون‌ها راستی‌آزمایی می‌شود تا غلطِ تایپیِ ستون بی‌صدا نپذیرفته شود
        header = _header_map(upload.content)
        for column, value in fields.items():
            label = str(column).strip()
            if label not in header:
                unknown_columns.append(label)
                continue
            overlay.setdefault(row_number, {})[label] = "" if value is None else str(value)
            applied += 1
    if unknown_columns:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "این ستون‌ها در فایل نیستند: " + "، ".join(sorted(set(unknown_columns))),
        )
    if not applied:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "اصلاحی اعمال نشد")

    structure["overlay"] = overlay
    upload.structure_json = json.dumps(structure, ensure_ascii=False)
    db = ctx.db
    db.add(upload)
    db.flush()

    summary = preview_summary(db, upload)
    return ToolOutcome(
        content=json_content({"patched": applied, **summary}),
        ui={"kind": "upload_report", "summary": {k: v for k, v in summary.items() if k != "rows"}},
        summary=f"{applied} سلول اصلاح شد؛ {summary['valid_count']} ردیف سالم، {summary['invalid_count']} خطادار",
    )


def _describe_patch(upload_id, edits=None, **_):
    count = len(edits or [])
    return f"اصلاح {count} ردیف فایل #{upload_id}"


patch_upload_rows.describe = _describe_patch


@tool(
    name="import_personnel",
    description=(
        "ورود ردیف‌های سالمِ اکسل پرسنل به سامانه — ساخت پرسنل، زنجیرهٔ ارزیابی و حساب‌ها. فقط پس از تأیید صریح کاربر "
        "اجرا می‌شود و پیش از اجرا، اعتبارسنجی از نو انجام می‌گیرد."
    ),
    category="فایل‌ها",
    read_only=False,
    risky=True,
    capabilities=(Capability.manage_personnel,),
    roles=(UserRole.hr,),
    parameters={
        "type": "object",
        "properties": {"upload_id": {"type": "integer"}},
        "required": ["upload_id"],
    },
)
def import_personnel(ctx: ToolContext, upload_id: int) -> ToolOutcome:
    """بدنهٔ این ابزار اجرا نمی‌کند؛ فقط می‌گوید «آمادهٔ تأیید هست».

    اجرای واقعی در `execute_import` است که تنها از نقطهٔ تأیید صدا زده
    می‌شود — و همان‌جا دوباره اعتبارسنجی و مجوز سنجیده می‌شود.
    """
    upload = _upload_or_404(ctx, upload_id)
    if not is_personnel_staging(upload):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "ورود گروهی فقط برای اکسل پرسنل است")
    structure = structure_of(upload)
    if structure.get("committed"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "این فایل قبلاً وارد سامانه شده است")
    preview = revalidate(ctx.db, upload)
    if preview.file_errors:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, preview.file_errors[0])
    if not preview.valid:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "هیچ ردیف سالمی برای ورود وجود ندارد")
    return ToolOutcome(
        content=json_content({
            "ready_for_confirmation": True,
            "valid_rows": len(preview.valid),
            "invalid_rows_will_be_skipped": len(preview.invalid),
            "accounts_to_create": sum(1 for row in preview.valid if row.username),
        }),
        summary=f"پیشنهاد ورود {len(preview.valid)} ردیف پرسنل از «{upload.filename}»",
    )


def _describe_import(upload_id, **_):
    return f"ورود ردیف‌های سالم فایل #{upload_id} به سامانه"


import_personnel.describe = _describe_import


def execute_import(ctx: ToolContext, upload_id: int) -> ToolOutcome:
    """اجرای واقعی ورود — فقط از نقطهٔ تأیید.

    فایل دوباره از صفر اعتبارسنجی می‌شود (همان قانونِ رابطِ دستی: بین
    پیش‌نمایش و ثبت، داده ممکن است عوض شده باشد) و درج با همان
    `commit_import` رخ می‌دهد که فرمِ دستی هم استفاده می‌کند.
    """
    from app.services.personnel_import import commit_import

    upload = _upload_or_404(ctx, upload_id)
    if not is_personnel_staging(upload):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "ورود گروهی فقط برای اکسل پرسنل است")
    structure = structure_of(upload)
    if structure.get("committed"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "این فایل قبلاً وارد سامانه شده است")
    preview = revalidate(ctx.db, upload)
    if preview.file_errors:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, preview.file_errors[0])
    if not preview.valid:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "هیچ ردیف سالمی برای ورود وجود ندارد")

    result = commit_import(ctx.db, preview, ctx.user.id)

    structure["committed"] = True
    structure["committed_result"] = {
        "created_personnel": result.created_personnel,
        "created_chains": result.created_chains,
        "skipped_rows": result.skipped_rows,
    }
    upload.structure_json = json.dumps(structure, ensure_ascii=False)
    ctx.db.add(upload)
    ctx.db.flush()

    accounts = [
        {
            "personnel_code": a.personnel_code,
            "full_name": a.full_name,
            "username": a.username,
            "temporary_password": a.temporary_password,
        }
        for a in result.accounts
    ]
    payload = {
        "imported": True,
        "created_personnel": result.created_personnel,
        "created_chains": result.created_chains,
        "skipped_rows": result.skipped_rows,
        "accounts": accounts,
    }
    if accounts:
        payload["note"] = "رمزهای موقت فقط همین یک‌بار نشان داده می‌شوند."
    return ToolOutcome(
        content=json_content(payload),
        ui={"kind": "import_result", "result": payload},
        summary=f"{result.created_personnel} پرسنل، {result.created_chains} زنجیره و {len(accounts)} حساب ساخته شد",
    )


def _header_map(content: bytes) -> dict[str, int]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    header = {}
    for cell in next(sheet.iter_rows(values_only=True), ()):
        label = str(cell).strip() if cell is not None else ""
        if label and label not in header:
            header[label] = len(header)
    workbook.close()
    return header




#: کنشِ دوپله‌ای: بدنهٔ `import_personnel` فقط «آمادگی و اعتبارسنجی» است؛
#: اجرای واقعی تنها از نقطهٔ تأیید، با همین تابع، رخ می‌دهد.
import_personnel.executor = execute_import
