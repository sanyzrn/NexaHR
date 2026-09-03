"""تکلیف کسی که از سازمان می‌رود.

تا امروز غیرفعال‌کردن یک پرسنل فقط جلوی *شروع* ارزیابی تازه را می‌گرفت. سه چیز
رها می‌ماند، و هر سه بی‌سروصدا هزینه داشتند:

* پروندهٔ باز همان‌جا معلق می‌ماند — در صف بررسی کسی، با یادآوری SLA فعال، برای
  کسی که دیگر در سازمان نیست.
* حساب کاربری‌اش باز می‌ماند؛ کسی که اخراج شده فردا هم می‌توانست وارد شود.
* علت خروج ثبت نمی‌شد. استعفا و اخراج و پایان قرارداد در گزارش HR یک چیز نیستند.
"""
from io import BytesIO

from app.core.clock import today_local
from app.models.enums import EvaluationStatus, PersonnelStatus, SeparationReason
from app.models.evaluation import EvaluationRecord
from app.models.personnel import Personnel
from tests.helpers import auth_header, make_access, make_personnel, make_user


def _hr(db_session):
    user = make_user(db_session, "hr")
    db_session.commit()
    return user


def _leave(client, hr, personnel, reason=SeparationReason.resignation, **extra):
    return client.patch(
        f"/api/personnel/{personnel.id}",
        json={"status": "inactive", "separation_reason": reason.value, **extra},
        headers=auth_header(hr),
    )


def test_a_reason_is_required(client, db_session):
    """رفتنِ بدون علت همان چیزی است که این تغییر برای حذفش آمد."""
    hr = _hr(db_session)
    personnel = make_personnel(db_session)
    db_session.commit()

    response = client.patch(
        f"/api/personnel/{personnel.id}",
        json={"status": "inactive"},
        headers=auth_header(hr),
    )
    assert response.status_code == 400, response.text
    assert "علت خروج" in response.json()["detail"]


def test_reason_and_date_are_recorded(client, db_session):
    hr = _hr(db_session)
    personnel = make_personnel(db_session)
    db_session.commit()

    response = _leave(client, hr, personnel, SeparationReason.dismissal)
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["separation_reason"] == "dismissal"
    # تاریخ داده نشده بود؛ امروز پیش‌فرض می‌شود تا ردیفی بدون تاریخ نماند.
    assert body["separation_date"] == today_local().isoformat()


def test_coming_back_clears_the_separation(client, db_session):
    """علتِ کهنه روی پروندهٔ یک نفرِ شاغل، بدترین نوع دادهٔ مانده است."""
    hr = _hr(db_session)
    personnel = make_personnel(db_session)
    db_session.commit()
    _leave(client, hr, personnel)

    back = client.patch(
        f"/api/personnel/{personnel.id}",
        json={"status": "active"},
        headers=auth_header(hr),
    )
    assert back.status_code == 200, back.text
    assert back.json()["separation_reason"] is None
    assert back.json()["separation_date"] is None


def test_the_open_case_is_cancelled_not_left_hanging(client, db_session):
    """پروندهٔ کسی که رفته ادامه‌پذیر نیست؛ معلق ماندنش فقط صف را کثیف می‌کند."""
    hr = _hr(db_session)
    supervisor = make_user(db_session, "unit_supervisor", capabilities=[])
    deputy = make_user(db_session, "deputy", capabilities=[])
    ceo = make_user(db_session, "ceo", capabilities=[])
    personnel = make_personnel(db_session)
    make_access(db_session, personnel, supervisor, deputy, ceo)
    db_session.commit()

    started = client.post(
        "/api/evaluations", json={"subject_personnel_id": personnel.id}, headers=auth_header(supervisor)
    )
    assert started.status_code == 201, started.text

    assert _leave(client, hr, personnel).status_code == 200

    record = db_session.get(EvaluationRecord, started.json()["id"])
    db_session.refresh(record)
    assert record.status is EvaluationStatus.cancelled


def test_the_account_can_no_longer_sign_in(client, db_session):
    """«غیرفعالش کردم» باید همان لحظه معنا داشته باشد، نه بعد از انقضای توکن."""
    hr = _hr(db_session)
    personnel = make_personnel(db_session)
    account = make_user(db_session, "employee", personnel_id=personnel.id, capabilities=[])
    db_session.commit()
    before_version = account.token_version

    assert _leave(client, hr, personnel).status_code == 200

    db_session.refresh(account)
    assert account.is_active is False
    # بالا رفتن نسخهٔ توکن یعنی نشست‌های زنده همان لحظه باطل می‌شوند.
    assert account.token_version > before_version
    assert client.get("/api/auth/me", headers=auth_header(account)).status_code == 401


def test_a_departure_does_not_disturb_anyone_else(client, db_session):
    """گاردِ کور: رفتنِ یک نفر نباید حساب کس دیگری را ببندد."""
    hr = _hr(db_session)
    leaver = make_personnel(db_session)
    stayer = make_personnel(db_session)
    stayer_account = make_user(
        db_session, "employee", personnel_id=stayer.id, capabilities=[]
    )
    db_session.commit()

    assert _leave(client, hr, leaver).status_code == 200

    db_session.refresh(stayer_account)
    assert stayer_account.is_active is True
    assert db_session.get(Personnel, stayer.id).status is PersonnelStatus.active


def test_the_reason_reaches_the_hr_export(client, db_session):
    """داده‌ای که فقط وارد می‌شود و هیچ‌جا بیرون نمی‌آید، ثبت نشده است.

    کل توجیه این ستون‌ها گزارش‌گیری بود — «نرخ استعفا در یک واحد یک سیگنال
    است». اگر در خروجی اکسل نباشند، آن سیگنال هیچ‌وقت خوانده نمی‌شود.
    """
    from openpyxl import load_workbook

    hr = _hr(db_session)
    personnel = make_personnel(db_session, full_name="کارمند رفته")
    db_session.commit()
    _leave(client, hr, personnel, SeparationReason.resignation)

    response = client.get("/api/personnel/export.xlsx", headers=auth_header(hr))
    assert response.status_code == 200, response.text

    sheet = load_workbook(BytesIO(response.content)).active
    header = [c.value for c in sheet[1]]
    assert "علت خروج" in header
    assert "تاریخ خروج" in header

    reason_at = header.index("علت خروج")
    rows = [r for r in sheet.iter_rows(min_row=2, values_only=True) if r[1] == "کارمند رفته"]
    assert rows and rows[0][reason_at] == "استعفا"
